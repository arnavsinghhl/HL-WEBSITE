import json
import sys
from pathlib import Path

from openpyxl import load_workbook


ALIASES = {
    "idNumber": {
        "idnumber",
        "id",
        "memberid",
        "distributorid",
        "supervisorid",
        "associateid",
        "customerid",
        "hlid",
        "hblid",
        "herbalifeid",
    },
    "name": {
        "name",
        "fullname",
        "membername",
        "distributorname",
        "supervisorname",
        "associatename",
        "customername",
    },
    "ppv": {"ppv", "personalvolume", "personalpv", "personalpointvolume"},
    "totalVolume": {"totalvolume", "totalvol", "volume", "totalpv", "tv", "pointvolume"},
    "currentMonth": {"currentmonth", "currentmonthname", "month", "monthname", "currentmnth"},
    "sponsorName": {"sponsorname", "sponsor", "sponsername", "sponser", "upline", "uplinename"},
    "volumeRequired": {
        "volumerequired",
        "requiredvolume",
        "volumerequiredforjimcorbettqualification",
        "jimcorbettrequiredvolume",
        "required",
        "shortfall",
        "balancevolume",
        "volumeshortfall",
        "neededvolume",
    },
    "result": {"result", "status", "qualificationresult", "progress", "progressstatus", "currentstatus"},
}


def normalize(value):
    return "".join(ch for ch in str(value or "").lower() if ch.isalnum())


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def header_matches(header, key):
    if not header:
        return False
    if header in ALIASES[key]:
        return True
    if key == "idNumber":
        return header.endswith("id") or "idnumber" in header or "distributorid" in header
    if key == "name":
        return "name" in header
    if key == "ppv":
        return "ppv" in header or "personalvolume" in header
    if key == "totalVolume":
        return ("total" in header and ("volume" in header or "pv" in header)) or header == "volume"
    if key == "currentMonth":
        return "month" in header
    if key == "sponsorName":
        return "sponsor" in header or "sponser" in header or "upline" in header
    if key == "volumeRequired":
        return (
            "required" in header
            or "shortfall" in header
            or "balancevolume" in header
            or "needed" in header
        )
    if key == "result":
        return "result" in header or "status" in header or "progress" in header
    return False


def find_index(headers, key):
    for index, header in enumerate(headers):
        if header_matches(header, key):
            return index
    return None


def find_header_row(rows):
    best = None
    for row_index, row in enumerate(rows[:30]):
        headers = [normalize(value) for value in row]
        indexes = {key: find_index(headers, key) for key in ALIASES}
        score = sum(index is not None for index in indexes.values())
        has_id = indexes["idNumber"] is not None
        if has_id and (best is None or score > best["score"]):
            best = {"row_index": row_index, "indexes": indexes, "score": score}
    return best


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: parse_excel.py <xlsx-file>")

    workbook_path = Path(sys.argv[1])
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if len(rows) < 2:
        print("[]")
        return

    header = find_header_row(rows)
    if not header:
        print("[]")
        return

    indexes = header["indexes"]
    records = []

    for row in rows[header["row_index"] + 1:]:
        def value_for(key):
            index = indexes[key]
            if index is None or index >= len(row):
                return ""
            return cell_text(row[index])

        record = {
            "idNumber": value_for("idNumber"),
            "name": value_for("name"),
            "ppv": value_for("ppv"),
            "totalVolume": value_for("totalVolume"),
            "currentMonth": value_for("currentMonth"),
            "sponsorName": value_for("sponsorName"),
            "volumeRequired": value_for("volumeRequired"),
            "result": value_for("result"),
        }

        if record["idNumber"]:
            records.append(record)

    print(json.dumps(records))


if __name__ == "__main__":
    main()
